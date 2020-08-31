import pylbc
import pandas as pd
from datetime import date

query = pylbc.Search()
query.set_square(15, 50)
query.set_price(300, 1000)
query.set_category('locations')
query.set_real_estate_types(['appartement'])
query.add_city('Thionville', '57100')
query.set_query('meuble gare')

i = 0
df = pd.DataFrame()

for result in query.iter_results():
    i = i + 1
    data = str(result).split('"')
    df[i] = data

df.drop(df.index[[0, 1, 2, 3, 4, 7, 9, 10, 11, 12]], inplace=True)
df = df.T
df.columns = ['Date', 'Prix', 'Surface']
df['Prix'] = df['Prix'].str[8:11]
df['Surface'] = df['Surface'].str[9:11]

df['Prix'] = df['Prix'].astype(float)
df['Surface'] = df['Surface'].astype(float)
df['Prix/m2'] = df['Prix'] / df['Surface']

df['Ancienneté'] = pd.to_datetime(date.today()) - pd.to_datetime(df['Date'])
df['Date'] = date.today()


def append_df_to_excel(filename, df, sheet_name='Data', startrow=None,
                       truncate_sheet=False,):

    from openpyxl import load_workbook

    writer = pd.ExcelWriter(filename, engine='openpyxl')

    try:
        writer.book = load_workbook(filename)

        if startrow is None and sheet_name in writer.book.sheetnames:
            startrow = writer.book[sheet_name].max_row

        if truncate_sheet and sheet_name in writer.book.sheetnames:
            idx = writer.book.sheetnames.index(sheet_name)
            writer.book.remove(writer.book.worksheets[idx])
            writer.book.create_sheet(sheet_name, idx)

        writer.sheets = {ws.title: ws for ws in writer.book.worksheets}
    except FileNotFoundError:
        pass

    df.to_excel(writer, sheet_name, startrow=startrow, index=None, header=None)

    writer.save()

print(df.mean(axis=0))
append_df_to_excel(r'/Users/maximethomas/Desktop/Data.xlsx', df)
